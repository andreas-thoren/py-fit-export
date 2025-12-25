import json
from collections.abc import Callable
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from garmin_fit_sdk import Decoder, Stream
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet


def _export_excel_wrapper(
    out_path: Path,
    ws_name: str,
    tbl_name: str,
    inner_func: Callable[[Worksheet, Table], None],
) -> None:
    wb = load_workbook(out_path)
    try:
        ws = wb[ws_name]
        tbl: Table = ws.tables[tbl_name]
        inner_func(ws, tbl)
        wb.save(out_path)
    finally:
        wb.close()


def export_activity_to_excel(
    activity_path: Path,
    column_map: dict[str, str],
    out_path: Path,
    ws_name: str,
    tbl_name: str,
) -> None:
    def _export_activity_inner(ws: Worksheet, tbl: Table) -> None:
        _excel_exporter(activity_path, column_map, ws, tbl)

    _export_excel_wrapper(out_path, ws_name, tbl_name, _export_activity_inner)


def export_activities_to_excel(
    activity_paths: list[Path],
    column_map: dict[str, str],
    out_path: Path,
    ws_name: str,
    tbl_name: str,
) -> None:
    def _export_activities_inner(ws: Worksheet, tbl: Table) -> None:
        for activity_path in activity_paths:
            _excel_exporter(activity_path, column_map, ws, tbl)

    _export_excel_wrapper(out_path, ws_name, tbl_name, _export_activities_inner)


def _excel_exporter(
    activity_path: Path, column_map: dict[str, str], ws: Worksheet, tbl: Table
) -> None:
    fit_info = extract_fit_info(activity_path)
    key_info = extract_key_info(fit_info)
    row_values: dict[str, Any] = {}
    for col_old, col_new in column_map.items():
        row_values[col_new] = key_info[col_old]

    append_table_values(ws, tbl, row_values)


def extract_fit_info(activity_path: Path) -> dict[str, Any]:
    stream = Stream.from_file(activity_path)
    decoder = Decoder(stream)
    fit_info, fit_errors = decoder.read()

    if fit_errors:
        print(fit_errors)

    return fit_info


def extract_key_info(fit_info: dict[str, Any]) -> dict[str, Any]:
    session = fit_info["session_mesgs"][0]

    workout_name = None
    if fit_info.get("workout_mesgs"):
        workout_name = fit_info["workout_mesgs"][0].get("wkt_name")

    start_time = session.get("start_time")

    return {
        "wrk_sport": session.get("sport"),
        "wrk_date": start_time.date() if isinstance(start_time, datetime) else None,
        "wrk_name": workout_name,
        "wrk_length": session.get("total_distance"),
        "wrk_load": session.get("training_load_peak"),
    }


def export_to_json(export_path: Path, info_dct: dict[str, Any]) -> None:
    safe_obj = _make_json_safe(info_dct)
    export_path.write_text(
        json.dumps(safe_obj, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def _make_json_safe(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {k: _make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_make_json_safe(v) for v in obj]
    if hasattr(obj, "isoformat"):  # datetime
        return obj.isoformat()
    if isinstance(obj, bytes):
        return obj.hex()
    return obj


def _make_ref(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    return (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )


def append_table_values(
    ws: Worksheet, table: Table, row_values: dict[str, Any]
) -> None:
    # 1. Raise early if tabel has totals row since function will not work!
    if table.totalsRowShown:
        raise RuntimeError(
            "append_table_values() cannot be used on tables with a totals row; "
            "this requires row insertion logic."
        )

    # 2. Construct initial required metadata about table and new row
    columns = {tc.name: i for i, tc in enumerate(table.tableColumns)}
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    row_i = max_row + 1
    column_index_range = range(min_col, max_col + 1)

    # 3. Error checks. Check row_values type and content. Also check new row is empty!
    if not isinstance(row_values, dict):
        raise TypeError(f"row_values has to be a dict not a {type(row_values)}!")

    faulty_keys = row_values.keys() - columns.keys()
    if faulty_keys:
        bad = ", ".join(sorted(map(str, faulty_keys)))
        raise ValueError(f"row_values contains invalid column names: {bad}")

    if any(ws.cell(row=row_i, column=c).value is not None for c in column_index_range):
        raise ValueError(
            f"Row {row_i} is not empty in the table area; would overwrite data!"
        )

    # 4. Set cell values for new row
    for col, val in row_values.items():
        col_i = min_col + columns[col]
        ws.cell(row=row_i, column=col_i, value=val)

    # 5. Fix formating of new row and its cells. Also expand formulas to new rows!
    if max_row > min_row:
        ws.row_dimensions[row_i].height = ws.row_dimensions[max_row].height

        for col_i in column_index_range:
            src = ws.cell(row=max_row, column=col_i)
            dst = ws.cell(row=row_i, column=col_i)

            if src.has_style:
                dst._style = copy(src._style)  # pylint: disable=protected-access
            if dst.value is None and (
                isinstance(src.value, str) and src.value.startswith("=")
            ):
                dst.value = Translator(
                    src.value, origin=src.coordinate
                ).translate_formula(dst.coordinate)

    # 6. Resize table to include new row
    new_ref = _make_ref(min_col, min_row, max_col, row_i)
    table.ref = new_ref

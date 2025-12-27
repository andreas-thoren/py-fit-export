from typing import Any
from datetime import datetime, time, timedelta, timezone
from openpyxl.utils import get_column_letter

CET = timezone(timedelta(hours=1))


def excel_safe_datetime(v):
    if isinstance(v, datetime) and v.tzinfo is not None:
        return v.astimezone(CET).replace(tzinfo=None)
    if isinstance(v, time) and v.tzinfo is not None:
        return v.replace(tzinfo=None)
    return v


def make_json_safe(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {k: make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]
    if hasattr(obj, "isoformat"):  # datetime
        return obj.isoformat()
    if isinstance(obj, bytes):
        return obj.hex()
    return obj


def make_ref(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    return (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )

import json
from collections.abc import Callable
from copy import copy
from datetime import datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet
from py_fit_export.fit_info_extractor import FitInfoExtractor
from py_fit_export.utils import make_json_safe, make_ref, excel_safe_datetime


class FitToExcelExporter:
    """
    Export Garmin FIT activity metadata to an existing Excel table.

    This class encapsulates all logic required to open an Excel workbook,
    locate a worksheet and table, optionally filter activities based on
    extracted metadata, and append new rows to the table.

    After instantiation, call one of the following methods to perform
    the actual export:
    - `export_activity_to_excel` to export a single FIT file
    - `export_activities_to_excel` to export multiple FIT files

    Instantiating this class alone does not modify the Excel workbook.
    """

    def __init__(
        self,
        excel_path: Path,
        ws_name: str,
        tbl_name: str,
        filter_map: dict[str, Any] | None = None,
    ) -> None:
        """
        Parameters
        ----------
        excel_path : Path
            Path to the target Excel workbook.
        ws_name : str
            Name of the worksheet containing the target table.
        tbl_name : str
            Name of the Excel table to append rows to.
        filter_map : dict[str, Any] | None, optional
            Optional filters applied to extracted FIT metadata before exporting.

            Each key in `filter_map` must correspond to a key produced by
            `FitInfoExtractor.extract()`.

            Values may be either:
            - a literal value, in which case the extracted value must be
              non-None and exactly equal (`==`) for the activity to be exported
            - a callable `predicate(value) -> bool`, which is invoked with the
              extracted value and must return True for the activity to be
              exported

            If any filter fails, the activity is skipped and no Excel row
            is appended.
        """
        self.path = excel_path
        self.ws_name = ws_name
        self.tbl_name = tbl_name
        self.filter_map = copy(filter_map) if filter_map else {}

    def _export_excel_wrapper(
        self,
        inner_func: Callable[[Worksheet, Table], None],
    ) -> None:
        wb = load_workbook(self.path)
        try:
            try:
                ws = wb[self.ws_name]
            except KeyError as exc:
                raise KeyError(
                    f"Worksheet {self.ws_name!r} not found in {self.path}"
                ) from exc
            tbl = ws.tables.get(self.tbl_name)
            if tbl is None:
                raise KeyError(
                    f"Table {self.tbl_name!r} not found on sheet {self.ws_name!r}"
                )
            inner_func(ws, tbl)
            wb.save(self.path)
        finally:
            wb.close()

    def export_activity_to_excel(
        self,
        activity_path: Path,
        column_map: dict[str, str],
    ) -> None:
        def _export_activity_inner(ws: Worksheet, tbl: Table) -> None:
            self._excel_exporter(activity_path, column_map, ws, tbl)

        self._export_excel_wrapper(_export_activity_inner)

    def export_activities_to_excel(
        self,
        activity_paths: list[Path],
        column_map: dict[str, str],
    ) -> None:
        def _export_activities_inner(ws: Worksheet, tbl: Table) -> None:
            for activity_path in activity_paths:
                self._excel_exporter(activity_path, column_map, ws, tbl)

        self._export_excel_wrapper(_export_activities_inner)

    def _excel_exporter(
        self, activity_path: Path, column_map: dict[str, str], ws: Worksheet, tbl: Table
    ) -> None:
        # Extract info from fit file by delegating to FitInfoExtractor
        extractor = FitInfoExtractor(activity_path)
        key_info = extractor.extract()

        # If filter_map exists check that values match otherwise return
        for key, filter_val in self.filter_map.items():
            fit_info_val = key_info.get(key)

            if callable(filter_val):
                passed_filter = filter_val(fit_info_val)
            else:
                passed_filter = fit_info_val is not None and fit_info_val == filter_val

            if not passed_filter:
                print(
                    f"Value in {activity_path} for {key}="
                    f"{fit_info_val!r} not matching required filter!"
                    "\nAborting excel export!"
                )
                return

        # Convert key names to align with column names in column_map
        row_values: dict[str, Any] = {}
        for col_old, col_new in column_map.items():
            try:
                row_values[col_new] = key_info[col_old]
            except KeyError as exc:
                raise KeyError(
                    f"Missing extracted key {col_old!r} for {activity_path}"
                ) from exc

        self.append_table_values(ws, tbl, row_values)

    @staticmethod
    def append_table_values(
        ws: Worksheet, table: Table, row_values: dict[str, Any]
    ) -> None:
        # 1. Raise early if tabel has totals row since function will not work!
        if table.totalsRowShown:
            raise RuntimeError(
                "append_table_values() cannot be used on tables with a totals row; "
                "this requires row insertion logic."
            )

        # 2. Construct initial required metadata about table and new row
        columns = {tc.name: i for i, tc in enumerate(table.tableColumns)}
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        row_i = max_row + 1
        column_index_range = range(min_col, max_col + 1)

        # 3. Error checks. Check row_values type and content. Also check new row is empty!
        if not isinstance(row_values, dict):
            raise TypeError(f"row_values has to be a dict not a {type(row_values)}!")

        faulty_keys = row_values.keys() - columns.keys()
        if faulty_keys:
            bad = ", ".join(sorted(map(str, faulty_keys)))
            raise ValueError(f"row_values contains invalid column names: {bad}")

        if any(
            ws.cell(row=row_i, column=c).value is not None for c in column_index_range
        ):
            raise ValueError(
                f"Row {row_i} is not empty in the table area; would overwrite data!"
            )

        # 4. Set cell values for new row
        for col, val in row_values.items():
            safe_val = (
                excel_safe_datetime(val) if isinstance(val, (datetime, time)) else val
            )
            col_i = min_col + columns[col]
            ws.cell(row=row_i, column=col_i, value=safe_val)

        # 5. Fix formating of new row and its cells. Also expand formulas to new rows!
        if max_row > min_row:
            ws.row_dimensions[row_i].height = ws.row_dimensions[max_row].height

            for col_i in column_index_range:
                src = ws.cell(row=max_row, column=col_i)
                dst = ws.cell(row=row_i, column=col_i)

                if src.has_style:
                    dst._style = copy(src._style)  # pylint: disable=protected-access
                if dst.value is None and (
                    isinstance(src.value, str) and src.value.startswith("=")
                ):
                    dst.value = Translator(
                        src.value, origin=src.coordinate
                    ).translate_formula(dst.coordinate)

        # 6. Resize table to include new row
        new_ref = make_ref(min_col, min_row, max_col, row_i)
        table.ref = new_ref


def export_activity_to_excel(
    activity: Path,
    excel_path: Path,
    column_map: dict[str, str],
    ws_name: str,
    tbl_name: str,
    filter_map: dict[str, Any] | None = None,
) -> None:
    exporter = FitToExcelExporter(excel_path, ws_name, tbl_name, filter_map)
    exporter.export_activity_to_excel(activity, column_map)


def export_activities_to_excel(
    activities: list[Path],
    excel_path: Path,
    column_map: dict[str, str],
    ws_name: str,
    tbl_name: str,
    filter_map: dict[str, Any] | None = None,
) -> None:
    exporter = FitToExcelExporter(excel_path, ws_name, tbl_name, filter_map)
    exporter.export_activities_to_excel(activities, column_map)


def export_to_json(export_path: Path, info_dct: dict[str, Any]) -> None:
    safe_obj = make_json_safe(info_dct)
    export_path.write_text(
        json.dumps(safe_obj, indent=2, ensure_ascii=False), encoding="utf-8"
    )
